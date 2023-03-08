import numpy as np
import scipy.signal as signal
from scipy.io import wavfile
import soundfile as sf
import openpyxl as px
from scipy.signal import firwin, filtfilt

#Varibles
Sheet = 'Band-Pass Filter'
dB = [0,0,10,10,20,20,30,30,40,40,50,50]
SNR_col = ['B','D','G','I','L','N','Q','S','V','X','AA','AC']
MSE_col = ['C','E','H','J','M','O','R','T','W','Y','AB','AD']
row = 4
ExcelFile = 'C:/Users/Matt/Documents/Project/CS-M/Experiments/FilterTests/FilterTestResults.xlsx'

# Upload to Excel
wb = px.load_workbook(ExcelFile)
wsbook = wb.active
ws = wb[Sheet]

for n in range(len(SNR_col)):
    for i in range(40):
        # Trim the Audio
        def trim_wav( originalWavPath, start, end ):
            sampleRate, waveData = wavfile.read( originalWavPath )
            startSample = int( start * sampleRate )
            endSample = int( end * sampleRate )
            wavfile.write( 'File.wav', sampleRate, waveData[startSample:endSample])

        # Trim Audio File to 10 sec
        if (n&2) == 0:
            Type = 'normal'
        else:
            Type = 'murmur'
        trim_wav("C:/Users/Matt/Documents/Project/CS-M/Experiments/FilterTests/"+Type+"/"+str(dB[n])+"dB/("+str(i+1)+").wav", 0,10)

        # Read in the audio file
        frames, sample_rate = sf.read('File.wav')

       # Band-pass Filter
        # Define filter parameters
        fmin = 20  # Lower cutoff frequency
        fmax = 250  # Upper cutoff frequency

        # Calculate the filter order
        nyquist_freq = 0.5 * sample_rate
        width = 100  # Width of the transition band, in Hz
        z = int(np.ceil(4 * nyquist_freq / width))

        # Design the bandpass filter
        b = firwin(z, [fmin, fmax], pass_zero=False, fs=sample_rate)

        # Apply the filter to the audio file
        filtered_frames = filtfilt(b, [1], frames)


        # Export new WAV file
        sf.write('Band-Pass_Filter.wav', filtered_frames, sample_rate)


        # Filter Effectiveness Tests
        # Calculate the Noise% (High is good)
        signal_power = np.sum(np.square(frames))
        noise_power = np.sum(np.square(frames - filtered_frames))
        Noise = noise_power / signal_power * 100
        print('Noise%: ',Noise)

        # Calculate the MSE (low is good)
        MSE = np.mean(np.square(frames - filtered_frames))
        print('MSE: ',MSE)


        # Upload to Excel
        print(n+1)
        ws[SNR_col[n]+str(row + i)] = Noise
        ws[MSE_col[n]+str(row + i)] = MSE

        wb.save(ExcelFile)