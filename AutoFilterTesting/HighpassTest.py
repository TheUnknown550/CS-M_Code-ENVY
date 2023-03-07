import numpy as np
import scipy.signal as signal
from scipy.io import wavfile
import soundfile as sf
import openpyxl as px

#Varibles
Sheet = 'High-Pass Filter'
Type = 'murmur'
dB = 10
SNR_col = 'I'
MSE_col = 'J'
row = 4
ExcelFile = 'C:/Users/Matt/Documents/Project/CS-M/Experiments/FilterTests/FilterTestResults.xlsx'

# Upload to Excel
wb = px.load_workbook(ExcelFile)
wsbook = wb.active
ws = wb[Sheet]

for i in range(40):
    # Trim the Audio
    def trim_wav( originalWavPath, start, end ):
        sampleRate, waveData = wavfile.read( originalWavPath )
        startSample = int( start * sampleRate )
        endSample = int( end * sampleRate )
        wavfile.write( 'File.wav', sampleRate, waveData[startSample:endSample])
    
    # Trim Audio File to 10 sec
    trim_wav("C:/Users/Matt/Documents/Project/CS-M/Experiments/FilterTests/"+Type+"/"+str(dB)+"dB/("+str(i+1)+").wav", 0,10)

    # Read in the audio file
    frames, sample_rate = sf.read('File.wav')

    # High-pass filter
    # Define the filter parameters
    cutoff_freq = 20 # Hz
    order = 4
    # Calculate the filter coefficients using a Butterworth filter
    nyquist_freq = 0.5 * sample_rate
    cutoff = cutoff_freq / nyquist_freq
    b, a = signal.butter(order, cutoff, btype='highpass')

    # Apply the filter to the audio frames
    filtered_frames = signal.filtfilt(b, a, frames)


    # Export new WAV file
    sf.write('High-Pass_Filter.wav', filtered_frames, sample_rate)


    # Filter Effectiveness Tests
    # Calculate the SNR (High is good)
    signal_power = np.sum(np.square(frames))
    noise_power = np.sum(np.square(frames - filtered_frames))
    SNR = 10 * np.log10(signal_power / noise_power)
    print('SNR: ',SNR)

    # Calculate the MSE (low is good)
    MSE = np.mean(np.square(frames - filtered_frames))
    print('MSE: ',MSE)


    # Upload to Excel
    ws[SNR_col+str(row + i)] = SNR
    ws[MSE_col+str(row + i)] = MSE

    wb.save(ExcelFile)