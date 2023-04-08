import numpy as np
import scipy.signal as signal
from scipy.io import wavfile
import soundfile as sf
import openpyxl as px
from scipy.signal import firwin, filtfilt
from pydub import AudioSegment
import time
import math


#Varibles
Sheet = 'BandPass'
dB = [0,0,10,10,20,20,30,30,40,40,50,50]
SNR_col = ['B','E','H','K','N','Q','S','V','X','AA','AC','AF','AI','AL','AO','AR','AU','AX','BA','BD','BG','BJ','BM','BP','BS','BV','BY','CB','CE','CH','CK','CN','CQ','CT','CW','CZ']
MSE_col = ['C','F','I','L','O','R','T','W','Y','AB','AD','AG','AJ','AM','AP','AS','AV','AY','BB','BE','BH','BK','BN','BQ','BT','BW','BZ','CC','CF','CI','CL','CO','CR','CU','CX','DA']
row = 4
ExcelFile = 'C:/Users/Matt/Documents/Project/CS-M/Experiments/NoiseCancelTests/Results.xlsx'

# Upload to Excel
wb = px.load_workbook(ExcelFile)
wsbook = wb.active
ws = wb[Sheet]

for n in range(6):
    for i in range(40):
        # Trim the Audio
        def trim_wav( originalWavPath, start, end, name ):
            sampleRate, waveData = wavfile.read( originalWavPath )
            startSample = int( start * sampleRate )
            endSample = int( end * sampleRate )
            wavfile.write( name, sampleRate, waveData[startSample:endSample])

        # Trim Audio File to 10 sec
        if n < 3:
            trim_wav("C:/Users/Matt/Documents/Project/CS-M/Experiments/NoiseCancelTests/ControlledVaribles/Heart/normal/("+str(i+1)+").wav", 0,10, 'C:/Users/Matt/Documents/Project/CS-M/Experiments/NoiseCancelTests/TestingAudios/Original.wav')
            trim_wav('C:/Users/Matt/Documents/Project/CS-M/Experiments/NoiseCancelTests/MixedSounds/Sound('+str(n+1)+')/normal/('+str(i+1)+').wav', 0,10, 'C:/Users/Matt/Documents/Project/CS-M/Experiments/NoiseCancelTests/TestingAudios/MixedFile.wav')
            Output = 'C:/Users/Matt/Documents/Project/CS-M/Experiments/NoiseCancelTests/ControlledVaribles/BandPass/normal/('+str(n+1)+'-'+str(i+1)+').wav'
        else:
            trim_wav("C:/Users/Matt/Documents/Project/CS-M/Experiments/NoiseCancelTests/ControlledVaribles/Heart/murmur/("+str(i+1)+").wav", 0,10, 'C:/Users/Matt/Documents/Project/CS-M/Experiments/NoiseCancelTests/TestingAudios/Original.wav')
            trim_wav('C:/Users/Matt/Documents/Project/CS-M/Experiments/NoiseCancelTests/MixedSounds/Sound('+str(n-2)+')/murmur/('+str(i+1)+').wav', 0,10, 'C:/Users/Matt/Documents/Project/CS-M/Experiments/NoiseCancelTests/TestingAudios/MixedFile.wav')
            Output = 'C:/Users/Matt/Documents/Project/CS-M/Experiments/NoiseCancelTests/ControlledVaribles/BandPass/murmur/('+str(n-2)+'-'+str(i+1)+').wav'
        
        # Read in the audio file
        Original, sample_rate = sf.read('C:/Users/Matt/Documents/Project/CS-M/Experiments/NoiseCancelTests/TestingAudios/Original.wav')
        frames, sampleRate = sf.read('C:/Users/Matt/Documents/Project/CS-M/Experiments/NoiseCancelTests/TestingAudios/MixedFile.wav')        


        # Band-pass Filter
        # Define filter parameters
        fmin = 20  # Lower cutoff frequency
        fmax = 250  # Upper cutoff frequency

        # Calculate the filter order
        nyquist_freq = 0.5 * sample_rate
        width = 100  # Width of the transition band, in Hz
        x = int(np.ceil(4 * nyquist_freq / width))

        # Design the bandpass filter
        b = firwin(x, [fmin, fmax], pass_zero=False, fs=sample_rate)

        # Apply the filter to the audio file
        filtered_frames = filtfilt(b, [1], frames, axis=0)


        # Export new WAV file
        sf.write(Output, filtered_frames, sample_rate)


        # Filter Effectiveness Tests
        # Calculate the Noise% (High is good)
        # Trim the ANC array to match the length of the Test array
        filtered_frames = filtered_frames[:len(Original)]
        # reshape Original to have the same shape as filtered_frames
        Original = np.reshape(Original, (Original.shape[0], 1))
        Original = np.tile(Original, (1, filtered_frames.shape[1]))
        
        # Calculate SNR
        signal_power = np.mean(np.square(Original))
        noise_power = np.mean(np.square(Original - filtered_frames))
        if noise_power == 0:
            SNR = float('inf')  # infinite SNR for zero noise power
        else:
            SNR = 10 * math.log10(signal_power / noise_power)
        print('SNR: ',SNR)

        # Calculate the MSE (low is good)
        RMSE = np.sqrt(np.mean(np.square(Original - filtered_frames)))
        print('RMSE: ',RMSE)
        
        # Upload to Excel
        print('Low: ',n+1)
        ws[SNR_col[n]+str(row + i)] = SNR
        ws[MSE_col[n]+str(row + i)] = RMSE

        wb.save(ExcelFile)